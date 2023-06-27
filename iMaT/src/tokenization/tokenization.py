"""
File tokenization module for MIDI files

This module provides a variety of functions for tokenizing MIDI files in a selected directory.
These MIDI files are then converted to sequences of tokens using one or more tokenizers.
The tokenized sequences are saved as CSV files in a subdirectory of the original directory.
A user interface allows the user to select the tokenizers and the folder to be tokenized.

Dependencies:
- os, re, ast, datetime
- pandas (pd)
- openpyxl
- tqdm
- miditok: Tokenizers for MIDI files
- CLI interface: cli.cli_menu_structure

Main function: corpus_tokenization
"""

import ast
import datetime
import glob
import os
import re
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from miditok import CPWord, MIDILike, MuMIDI, Octuple, OctupleMono, REMI, REMIPlus, Structured, TSD
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from tqdm import tqdm

from src.cli.cli_menu_structures import display_menu_print_results, display_menu_print_textblock, \
    display_menu_request_selection

tokenizers_list = [
    ['REMI', REMI, '<One-Dimensional: Event-based, includes timing, bar info>'],
    ['REMIPlus', REMIPlus, '<One-Dimensional: Extended REMI, multi-track, multi-signature>'],
    ['MIDILike', MIDILike, '<One-Dimensional: Converts MIDI messages to tokens>'],
    ['TSD', TSD, '<One-Dimensional: Similar to MIDI-Like, uses explicit Duration tokens>'],
    ['Structured', Structured, '<One-Dimensional: Similar to TSD, consistent token type succession>'],
    ['CPWord', CPWord, '<Two-Dimensional: Uses embedding pooling to reduce sequence length>'],
    ['Octuple', Octuple, '<Two-Dimensional: Embedding pooling, single note representation>'],
    ['OctupleMono', OctupleMono, '<Two-Dimensional: Like Octuple, but suited for one track>'],
    ['MuMIDI', MuMIDI, '<Two-Dimensional: Multitrack tasks, uses embedding pooling>']
]

tokenizer_headers = {
    'REMI': ['Tokens'],
    'REMIPlus': ['Tokens'],
    'MIDILike': ['Tokens'],
    'TSD': ['Tokens'],
    'Structured': ['Tokens'],
    'CPWord': ['Family', 'Position', 'Pitch', 'Velocity', 'Duration', 'Rest'],
    'Octuple': ['Pitch', 'Velocity', 'Duration', 'Program', 'Position', 'Bar'],
    'OctupleMono': ['Pitch', 'Velocity', 'Duration', 'Position', 'Bar'],
    'MuMIDI': ['Type*', 'BarPosEnc', 'PositionPosEnc', 'Velocity', 'Duration']
}

tokenizer_additional_tokens = {
    'Chord': False,
    'Program': False,
    'Rest': True,
    'Tempo': False,
    'TimeSignature': True,
    'chord_maps': {
        '7aug': (0, 4, 8, 11),
        '7dim': (0, 3, 6, 9),
        '7dom': (0, 4, 7, 10),
        '7halfdim': (0, 3, 6, 10),
        '7maj': (0, 4, 7, 11),
        '7min': (0, 3, 7, 10),
        '9maj': (0, 4, 7, 10, 14),
        '9min': (0, 4, 7, 10, 13),
        'aug': (0, 4, 8),
        'dim': (0, 3, 6),
        'maj': (0, 4, 7),
        'min': (0, 3, 7),
        'sus2': (0, 2, 7),
        'sus4': (0, 5, 7)
    },
    'chord_tokens_with_root_note': True,
    'chord_unknown': False,
    'nb_tempos': 32,
    'programs': list(range(-1, 128)),
    'rest_range': (2, 16),
    'tempo_range': (40, 250),
    'time_signature_range': (8, 2)
}


def corpus_tokenization():
    """
    Main function for tokenizing MIDI files.

    It asks the user to select a directory and one or more tokenizers, and then tokenizes
    all MIDI files in the selected directory using the chosen tokenizers. The tokenized
    sequences are saved as CSV files in a subdirectory of the original directory.
    Tokenization results and errors are logged in an Excel file in the same subdirectory.

    Returns
    -------
    None
    """
    while True:
        folder_path = select_folder()
        if folder_path is None:
            return
        tokenizable_files = list_tokenizable_files(folder_path)
        tokenizable_files_display = tokenizable_files[:30]  # Only take the first 30 files for display
        if len(tokenizable_files) > 30:  # If there are more than 30 files, add an indicator at the end
            tokenizable_files_display.append("... (more files not shown)")

        music_files_dict = {
            "menu_displayed_text":
                [f"File Tokenization: Found MIDI files ({len(tokenizable_files)} files found)",
                 f"In Folder: '{folder_path}'",
                 "<To continue, please press Enter (enter 'r' to refresh)> ",
                 ["File Path"]],
            "menu_entries_results":
                [[file] for file in tokenizable_files_display]
        }
        refresh_choice = display_menu_print_results(music_files_dict)
        if refresh_choice.lower() != 'r':
            break

    selected_tokenizers = select_tokenizer()

    tokenization_dir = os.path.join(folder_path, "tokenized_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.mkdir(tokenization_dir)

    workbook = Workbook()
    workbook.save(tokenization_dir + r'\tokenization_log.xlsx')

    num_files_tokenized = len(tokenizable_files) * len(selected_tokenizers)
    num_tokenized = 0
    files_status = []

    for tokenizer_class in selected_tokenizers:
        # Look up the name of the tokenizer
        tokenizer_name = next((x[0] for x in tokenizers_list if x[1] == tokenizer_class), None)
        if not tokenizer_name:
            print(f"Error: Could not find tokenizer '{tokenizer_name}'")
            continue  # Skip if the tokenizer name is not found

        if tokenizer_name not in ['Structured', 'REMIPlus']:
            tokenizer = tokenizer_class(additional_tokens=tokenizer_additional_tokens, beat_res = {(0, 16): 16})
        else:
            tokenizer = tokenizer_class()

        tokenized_file_folder = os.path.join(tokenization_dir, f"tokenizer_{tokenizer_name}")
        os.mkdir(tokenized_file_folder)
        for file in tqdm(tokenizable_files, ncols=70):  # Wrap tokenizable_files with tqdm for progress bar
            try:
                midi_file_path = os.path.join(folder_path, file)
                tokens = tokenizer(midi_file_path)
                tokens_string = str(tokens)

                # check if the tokens are in a nested list
                if "[[" in tokens_string and "]]" in tokens_string:
                    tokens_list = extract_tokens_nested_list(tokens_string)
                else:
                    tokens_list = extract_tokens(tokens_string)

                # Convert the list of tokens to a DataFrame
                df = pd.DataFrame(tokens_list)

                # Get the column names for the current tokenizer
                headers = tokenizer_headers.get(tokenizer_name, [])

                # Check if the DataFrame has the right number of columns
                if len(df.columns) == len(headers):
                    df.columns = headers
                else:
                    print(
                        f'\nWarning: Number of columns in the DataFrame does not match the number of headers for '
                        f'tokenizer {tokenizer_name}. Defaulting to generic headers.\n')
                    df.columns = ['column' + str(i) for i in range(1, len(df.columns) + 1)]

                # Save the DataFrame to a CSV file
                base_name = os.path.splitext(file)[0]  # Get file base name without extension
                file_name_token = f"{base_name}_tokenizer_{tokenizer_name}_tokens.csv"
                file_path = os.path.join(tokenized_file_folder, file_name_token)

                df.to_csv(file_path, index=False)
                num_tokenized += 1
                files_status.append([tokenizer_name, file, "<successfully converted>"])

            except Exception as e:
                files_status.append([tokenizer_name, file, str(e)])
                exception_list = [tokenizer_name, os.path.join(tokenized_file_folder, file), str(e)]
                create_log_entry(exception_list, tokenization_dir + r'\tokenization_log.xlsx')

        combine_csv_files_in_directory(tokenized_file_folder, f"00_combined_tokenizer_{tokenizer_name}_tokens.csv")

    show_success_rate(files_status, num_files_tokenized, num_tokenized)


def show_success_rate(files_status, num_files_tokenized, num_tokenized):
    # Create text_dict with both tokenization success rate and failed files
    failed_files_status = [status for status in files_status if "<successfully converted>" not in status]
    if failed_files_status:  # If there are failed tokenizations
        # Filter out only the failed tokenizations and create text_dict with failed files only
        text_dict = {
            "menu_displayed_text": [
                "File Tokenization: Tokenization Summary",
                f"Conversion Success Rate {num_tokenized}/{num_files_tokenized} ({(num_tokenized / num_files_tokenized) * 100:.2f}%)",
                "<To continue, please press Enter>",
                ["Tokenizer name", "File name", "Details"]
            ],
            "menu_entries_results": [
                [str(tokenizer_name), str(file), f"'{str(status)}'"] for tokenizer_name, file, status in
                failed_files_status]
        }
    else:  # If there are no failed tokenizations
        # Show the first 30 successful tokenizations, or all if less than 30
        num_successful_files = min(30, len(files_status))
        successful_files_status = files_status[:num_successful_files]
        successful_files_display = successful_files_status
        if len(files_status) > 30:  # If there are more than 30 files, add an indicator at the end
            successful_files_display.append(["...", "...", "<more files not shown>"])
        text_dict = {
            "menu_displayed_text": [
                "File Tokenization: Tokenization Summary",
                f"Conversion Success Rate {num_tokenized}/{num_files_tokenized} ({(num_tokenized / num_files_tokenized) * 100:.2f}%)",
                "<To continue, please press Enter>",
                ["Tokenizer name", "File name", "Details"]
            ],
            "menu_entries_results": [
                [str(tokenizer_name), str(file), f"'{str(status)}'"] for tokenizer_name, file, status in
                successful_files_display]
        }
    display_menu_print_results(text_dict)


def combine_csv_files_in_directory(directory_path, output_file_name):
    """
    Combine all CSV files in a given directory into a single CSV file.

    Parameters
    ----------
    directory_path : str
        Path to the directory with the CSV files.
    output_file_name : str
        Name of the combined output CSV file.

    Returns
    -------
    None
    """

    # Get a list of all CSV files in the directory
    csv_files = glob.glob(os.path.join(directory_path, "*.csv"))

    # Initialize an empty list to hold DataFrames
    df_list = []
    # Loop through each CSV file
    for csv_file in tqdm(csv_files, ncols=70):  # Wrap csv_files with tqdm for progress bar
        # Read the CSV file into a DataFrame
        df = pd.read_csv(csv_file)

        # Extract the original file name from the CSV file name
        original_file_name = os.path.basename(csv_file).split('_tokenizer')[0]
        # Assuming file names are in the format "<original_name>_tokenizer_<tokenizer_name>_tokens.csv"

        # Insert it into a new column at the start of the DataFrame
        df.insert(0, 'filename', original_file_name)

        # Append the DataFrame to the list
        df_list.append(df)

    # Concatenate all DataFrames in the list
    combined_df = pd.concat(df_list, ignore_index=True)

    # Save the combined DataFrame to a CSV file
    combined_df.to_csv(os.path.join(directory_path, output_file_name), index=False)


def list_tokenizable_files(folder_path):
    """
    Lists all MIDI files in a given folder.

    Parameters
    ----------
    folder_path : str
        Path to the folder to be searched.

    Returns
    -------
    list
        List of MIDI file names.
    """
    tokenizable_extensions = ['.midi', '.mid']
    tokenizable_files = [f for f in os.listdir(folder_path) if os.path.splitext(f)[1].lower() in tokenizable_extensions]
    return tokenizable_files


def select_tokenizer():
    """
    Provides a user interface for tokenizer selection.

    The user can select one or more tokenizers from a predefined list.

    Returns
    -------
    list
        List of selected tokenizer classes.
    """
    tokenizer_dict = {
        "menu_displayed_text": [
            "File Tokenization: Tokenizer Selection",
            "Please select one of the following tokenizers by entering the corresponding index number:",
            "Which tokenizer do you want to select? (<No. of menu item>): ",
            ["Tokenizer", "Description"]
        ],
        "menu_entries": [[tokenizer[0], tokenizer[1], tokenizer[2]] for tokenizer in tokenizers_list] + [
            ["All", "return_all", "Apply all tokenizers"]]
    }
    choice = display_menu_request_selection(tokenizer_dict)

    if choice == "return_all":
        return [tokenizer[1] for tokenizer in tokenizers_list]
    else:
        return [choice]

def select_folder():
    """
    User interface for folder selection. The user can input the path to the desired folder.
    If the path does not exist or does not lead to a directory, an error message is displayed.

    Returns
    -------
    str
        Path to the selected folder.
    """
    parsable_extensions = ['.midi', '.mid']

    while True:
        textblock_dict = {
            "menu_displayed_text": [
                "-- Folder Selection --",
                "Please read the following information:",
                "<To continue and select a folder, please press Enter>",
                ["", "Information"],
            ],
            "menu_entries_text": [
                ["Parsable Music File Types",
                 ", ".join(parsable_extensions)],
                ["Folder Requirements",
                 "The folder must contain at least one parsable music file of the above types to be processed by this tool."],
                ["Return to Main Menu",
                 "If you do not have a suitable folder with valid files at the moment, press enter and close the selection display that will be displayed right after.\n"
                 "You will then be returned to the Main Menu."]
            ]
        }

        display_menu_print_textblock(textblock_dict)

        root = tk.Tk()
        # root.withdraw()  # Hide the root window

        # Open a file dialog and get the selected file path
        folder_path = filedialog.askdirectory()

        root.destroy()  # Destroy the root window

        if folder_path == "":  # If the user canceled the dialog
            return None

        if len(list_tokenizable_files(folder_path)) == 0:
            error_message_dict = {
                "menu_displayed_text": [
                    "File Conversion: Folder Selection - Error",
                    "The selected directory contains no parsable music files:",
                    "<To continue, please press Enter>",
                    ["", "Troubleshooting assistance:"]
                ],
                "menu_entries_text": [
                    ["Parsable Music File Types",
                     ", ".join(parsable_extensions)],
                    ["Reason 1:", "There are no parsable music files in the selected directory."],
                    ["Reason 2:", "Ensure the directory you select contains the music files you want to convert."],
                    ["Return to Main Menu",
                    "If you do not have a suitable folder with valid files at the moment, press enter and close the "
                    "selection display that will be displayed right after.\n"
                    "You will then be returned to the Main Menu."]
                ]
            }
            display_menu_print_textblock(error_message_dict)
            continue
        else:
            return folder_path


def extract_tokens(tokens_string):
    """
    Extracts tokens from a string representation of a list.

    Parameters
    ----------
    tokens_string : str
        String representation of a list of tokens.

    Returns
    -------
    list
        List of tokens.
    """
    # cut away the beginning of the string
    start_index = tokens_string.index('=[') + 2
    trimmed_string = tokens_string[start_index:]

    # cut away everything after the first appearing "]"
    end_index = trimmed_string.index(']')
    trimmed_string = trimmed_string[:end_index]

    # extract strings using regular expression
    search_pattern = r"'([^']*)'"
    token_list = re.findall(search_pattern, trimmed_string)

    return token_list


def extract_tokens_nested_list(tokens_string):
    """
    Extracts tokens from a string representation of a nested list.

    Parameters
    ----------
    tokens_string : str
        String representation of a nested list of tokens.

    Returns
    -------
    list
        List of token lists.
    """
    # cut away everything before the string "[TokSequence(tokens=["
    start_index = tokens_string.index('=[') + 2
    trimmed_string = tokens_string[start_index:]

    # cut away everything after the string ", ids="
    end_index = trimmed_string.index('], ids=')
    trimmed_string = trimmed_string[:end_index]

    search_pattern = r'\[([^\[\]]*?)\]'
    matched_lists = re.findall(search_pattern, trimmed_string)

    token_lists = [ast.literal_eval(selected_list) for selected_list in matched_lists]

    return token_lists


def tokenize_midi_file(midi_file_path, tokenizer_class, tokenized_file_folder):
    """
    Tokenizes a MIDI file using a specified tokenizer class and saves the tokenized sequence to a
    CSV file.

    Parameters
    ----------
    midi_file_path : str
        Path to the MIDI file to be tokenized.
    tokenizer_class : class
        Class of the tokenizer to be used.
    tokenized_file_folder : str
        Path to the folder where the CSV file will be saved.

    Returns
    -------
    None
    """
    # create the tokenizer and convert the MIDI to tokens
    tokenizer = tokenizer_class()
    tokens = tokenizer(midi_file_path)

    tokens_string = str(tokens)

    # check if the tokens are in a nested list
    if "[[" in tokens_string and "]]" in tokens_string:
        tokens_list = extract_tokens_nested_list(tokens_string)
    else:
        tokens_list = extract_tokens(tokens_string)

    # convert the list of tokens to a DataFrame
    df = pd.DataFrame(tokens_list)

    # save the DataFrame to a CSV file
    file_name_token = f"{tokenizer_class.__name__}_tokens.csv"
    file_path = os.path.join(tokenized_file_folder, file_name_token)

    df.to_csv(file_path, index=False, header=False)

    # print(f"Tokens saved to {file_path}")


def create_log_entry(log_list, log_list_path):
    """
    Appends a new entry to an existing .xlsx file.

    If the file does not exist, it is created.

    Parameters
    ----------
    log_list : list
        List of strings to be appended as a new row.
    log_list_path : str
        Path to the .xlsx file.

    Returns
    -------
    None
    """
    wb = load_workbook(log_list_path)
    ws = wb.active
    ws.append(log_list)
    wb.save(log_list_path)
