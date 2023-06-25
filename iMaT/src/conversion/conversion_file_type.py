import os
import os.path
import tkinter as tk
from datetime import datetime
from tkinter import filedialog

from music21 import converter
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from src.cli.cli_menu_structure import display_menu_print_results, display_menu_print_textblock, \
    display_menu_request_selection


def convert_multiple_files_filetype():
    """
    Main function for the conversion of music files found in a pre-defined folder.
    The function lists parsable music files, allows the user to select a conversion format,
    and attempts conversion. Successfully converted files are saved in a newly created directory.
    Failures during parsing and conversion are logged to respective .xlsx files.

    Returns
    -------
    None
    """
    while True:
        folder_path = select_folder()
        if folder_path is None:
            return
        parsable_files = list_parsable_files(folder_path)
        parsable_files_display = parsable_files[:30]  # Only take the first 30 files for display
        if len(parsable_files) > 30:  # If there are more than 30 files, add an indicator at the end
            parsable_files_display.append("... (more files not shown)")

        music_files_dict = {
            "menu_displayed_text":
                ["File Conversion: Found music files",
                 f"In Folder: '{folder_path}'  ({len(parsable_files)} files found)",
                 "<To continue, please press Enter (enter 'r' to refresh)> ",
                 ["File Path"]],
            "menu_entries_results":
                [[file] for file in parsable_files_display]
        }
        refresh_choice = display_menu_print_results(music_files_dict)
        if refresh_choice.lower() != 'r':
            break

    conversion_format = select_conversion_format()

    conversion_dir = os.path.join(folder_path, "converted_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.mkdir(conversion_dir)

    workbook1 = Workbook()
    workbook1.save(conversion_dir + r'\parsing_log.xlsx')

    workbook2 = Workbook()
    workbook2.save(conversion_dir + r'\conversion_log.xlsx')

    num_files = len(parsable_files)
    num_converted = 0
    files_status = []

    print("\nConverting files...")
    for file in tqdm(parsable_files, ncols=70):  # Wrap parsable_files with tqdm for progress bar
        try:
            file_path = os.path.join(folder_path, file)
            score = converter.parse(file_path)
            try:
                converted_file_path = os.path.join(conversion_dir,
                                                   os.path.splitext(os.path.basename(file))[0] + conversion_format)
                fp = score.write(conversion_format, fp=converted_file_path)
                num_converted += 1
                files_status.append([file, "<successfully converted>"])
            except Exception as e:
                files_status.append([file, str(e)])
                list = [file_path, str(e)]
                create_log_entry(list, conversion_dir + '\conversion_log.xlsx')
        except Exception as e:
            files_status.append([file, str(e)])
            list = [file_path, str(e)]
            create_log_entry(list, conversion_dir + '\parsing_log.xlsx')

    # Create text_dict with both conversion success rate and failed files
    text_dict = {
        "menu_displayed_text": [
            "File Conversion: Conversion Summary",
            f"Conversion Success Rate {num_converted}/{num_files} ({(num_converted / num_files) * 100:.2f}%)",
            "<To continue, please press Enter>",
            ["File name", "Details"]
        ],
        "menu_entries_results": [
            [str(file), f"'{str(status)}'"] for file, status in files_status]
    }

    display_menu_print_results(text_dict)


def list_parsable_files(folder_path):
    """
    Lists all parsable music files in a given folder.

    Parameters
    ----------
    folder_path : str
        Path to the folder to be searched.

    Returns
    -------
    list
        List of parsable music file names.
    """
    parsable_extensions = ['.abc', '.capx', '.gex', '.humdrum', '.krn', '.mei', '.midi', '.mid', '.musedata',
                           '.musicxml', '.mxl', '.noteworthy', '.nwc', '.romanText', '.rntxt', '.scala',
                           '.tinynotation', '.volpiano']
    parsable_files = [f for f in os.listdir(folder_path) if os.path.splitext(f)[1].lower() in parsable_extensions]
    return parsable_files


def select_folder():
    """
    User interface for folder selection. The user can input the path to the desired folder.
    If the path does not exist or does not lead to a directory, an error message is displayed.

    Returns
    -------
    str
        Path to the selected folder.
    """
    parsable_extensions = ['.abc', '.capx', '.gex', '.humdrum', '.krn', '.mei', '.midi', '.mid', '.musedata',
                           '.musicxml', '.mxl', '.noteworthy', '.nwc', '.romanText', '.rntxt', '.scala',
                           '.tinynotation', '.volpiano']

    while True:
        textblock_dict = {
            "menu_displayed_text": [
                "-- Folder Selection --",
                "Please read the following information:",
                "<To continue and select a folder, please press Enter>",
                ["", "Information"],
            ],
            "menu_entries_text": [
                ["Parsable Music File Types",
                 ", ".join(parsable_extensions)],
                ["Folder Requirements",
                 "The folder must contain at least one parsable music file of the above types to be processed by this tool."],
                ["Return to Main Menu",
                 "If you do not have a suitable folder with valid files at the moment, press enter and close the selection display that will be displayed right after.\n"
                 "You will then be returned to the Main Menu."]
            ]
        }

        display_menu_print_textblock(textblock_dict)

        root = tk.Tk()
        # root.withdraw()  # Hide the root window

        # Open a file dialog and get the selected file path
        folder_path = filedialog.askdirectory()

        root.destroy()  # Destroy the root window

        if folder_path is "":  # If the user canceled the dialog
            return None

        parsable_files = list_parsable_files(folder_path)
        if len(parsable_files) == 0:
            error_message_dict = {
                "menu_displayed_text": [
                    "File Conversion: Folder Selection - Error",
                    "The selected directory contains no parsable music files:",
                    "<To continue, please press Enter>",
                    ["", "Troubleshooting assistance:"]
                ],
                "menu_entries_text": [
                    ["Parsable Music File Types",
                     ", ".join(parsable_extensions)],
                    ["Reason 1:", "There are no parsable music files in the selected directory."],
                    ["Reason 2:", "Ensure the directory you select contains the music files you want to convert."],
                    ["Return to Main Menu",
                    "If you do not have a suitable folder with valid files at the moment, press enter and close the "
                    "selection display that will be displayed right after.\n"
                    "You will then be returned to the Main Menu."]
                ]
            }
            display_menu_print_textblock(error_message_dict)
            continue
        else:
            return folder_path


def select_conversion_format():
    """
    User interface for conversion format selection. The user can choose a conversion format from a list.
    The function loops until a valid choice is made.

    Returns
    -------
    str
        Chosen conversion format.
    """
    # Get all available formats.
    formats = [['.midi', '.mid', '.midi/.mid ok'],
               ['.mxl', '.mxl', 'ok'],
               ['.musicxml', '.musicxml', 'ok'],
               ['.noteworthy', '.noteworthy', 'Attention: a bytes-like object is required'],
               ['.nwc', '.nwc', 'Attention: a bytes-like object is required'],
               ['.scala', '.scala', 'ok'],
               ['.tinynotation', '.tinynotation', 'ok'],
               ['.abc', '.abc', 'ok'],
               ['.capx', '.capx', 'ok'],
               ['.gex', '.gex', 'conversion not supported'],
               ['.humdrum', '.humdrum', 'ok'],
               ['.krn', '.krn', 'ok'],
               ['.mei', '.mei', 'MEI export is not yet implemented.'],
               ['.musedata', '.musedata', 'ok'],
               ['.romanText', '.romanText', 'ok'],
               ['.rntxt', '.rntxt', 'ok'],
               ['.volpiano', '.volpiano', 'ok']]

    # Loop to allow re-selection in case of invalid choice.
    while True:
        # Construct data container for displaying the menu.
        imat_data_container = {
            "menu_displayed_text": [
                "File Conversion: Conversion formats",
                "Please select one of the following conversion formats by entering the corresponding index number:",
                "Which format do you want to select? (<No. of menu item>): ",
                ["Format"]
            ],
            "menu_entries": formats
        }

        # Display menu and return the user's choice.
        return display_menu_request_selection(imat_data_container)


def create_log_entry(list, list_path):
    """
    Appends a new entry to an existing .xlsx file. If the file does not exist, it is created.

    Parameters
    ----------
    list : list
        List of strings to be appended as a new row.
    list_path : str
        Path to the .xlsx file.

    Returns
    -------
    None
    """
    wb = load_workbook(list_path)
    ws = wb.active
    ws.append(list)
    wb.save(list_path)

# convert_multiple_files_filetype()
