"""
File tokenization module for MIDI files

This module provides a variety of functions for tokenizing MIDI files in a selected directory.
These MIDI files are then converted to sequences of tokens using one or more tokenizers.
The tokenized sequences are saved as CSV files in a subdirectory of the original directory.
A user interface allows the user to select the tokenizers and the folder to be tokenized.

Dependencies:
- os, re, ast, datetime
- pandas (pd)
- openpyxl
- tqdm
- miditok: Tokenizers for MIDI files
- CLI interface: src.cli.cli_menu_structure

Main function: corpus_tokenization
"""

import ast
import datetime
import os
import re

import pandas as pd
import glob
from miditok import REMI, REMIPlus, MIDILike, TSD, Structured, CPWord, Octuple, OctupleMono, MuMIDI
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from tqdm import tqdm

from src.cli.cli_menu_structure import display_menu_print_results, display_menu_request_selection, \
    display_menu_print_textblock

tokenizers_list = [
    ['REMI', REMI, '<One-Dimensional: Event-based, includes timing, bar info>'],
    ['REMIPlus', REMIPlus, '<One-Dimensional: Extended REMI, multi-track, multi-signature>'],
    ['MIDILike', MIDILike, '<One-Dimensional: Converts MIDI messages to tokens>'],
    ['TSD', TSD, '<One-Dimensional: Similar to MIDI-Like, uses explicit Duration tokens>'],
    ['Structured', Structured, '<One-Dimensional: Similar to TSD, consistent token type succession>'],
    ['CPWord', CPWord, '<Two-Dimensional: Uses embedding pooling to reduce sequence length>'],
    ['Octuple', Octuple, '<Two-Dimensional: Embedding pooling, single note representation>'],
    ['OctupleMono', OctupleMono, '<Two-Dimensional: Like Octuple, but suited for one track>'],
    ['MuMIDI', MuMIDI, '<Two-Dimensional: Multitrack tasks, uses embedding pooling>']
]

tokenizer_headers = {
    'REMI': ['Tokens'],
    'REMIPlus': ['Tokens'],
    'MIDILike': ['Tokens'],
    'TSD': ['Tokens'],
    'Structured': ['Tokens'],
    'CPWord': ['Family', 'Bar/Position', 'Pitch', 'Velocity', 'Duration', 'None'],
    'Octuple': ['Pitch', 'Velocity', 'Duration', 'Program (track)', 'Position', 'Bar'],
    'OctupleMono': ['Pitch', 'Velocity', 'Duration', 'Position', 'Bar'],
    'MuMIDI': ['Type*', 'BarPosEnc', 'BarPosEnc', 'Velocity', 'Duration']
}


def corpus_tokenization():
    """
    Main function for tokenizing MIDI files.

    It asks the user to select a directory and one or more tokenizers, and then tokenizes
    all MIDI files in the selected directory using the chosen tokenizers. The tokenized
    sequences are saved as CSV files in a subdirectory of the original directory.
    Tokenization results and errors are logged in an Excel file in the same subdirectory.

    Returns
    -------
    None
    """
    while True:
        folder_path = select_folder()
        if folder_path is not None:
            break

    while True:
        tokenizable_files = list_tokenizable_files(folder_path)
        tokenizable_files_display = tokenizable_files[:30]  # Only take the first 30 files for display
        if len(tokenizable_files) > 30:  # If there are more than 30 files, add an indicator at the end
            tokenizable_files_display.append("... (more files not shown)")

        music_files_dict = {
            "menu_displayed_text":
                ["File Tokenization: Found MIDI files",
                 f"In Folder: '{folder_path}'  ({len(tokenizable_files)} files found)",
                 "<To continue, please press Enter (enter 'r' to refresh)> ",
                 ["File Path"]],
            "menu_entries_results":
                [[file] for file in tokenizable_files_display]
        }
        refresh_choice = display_menu_print_results(music_files_dict)
        if refresh_choice.lower() != 'r':
            break

    selected_tokenizers = select_tokenizer()

    tokenization_dir = os.path.join(folder_path, "tokenized_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.mkdir(tokenization_dir)

    workbook = Workbook()
    workbook.save(tokenization_dir + r'\tokenization_log.xlsx')

    num_files = len(tokenizable_files)
    num_files_tokenized = len(tokenizable_files) * len(selected_tokenizers)
    num_tokenized = 0
    files_status = []

    print("\nTokenizing files...")
    for tokenizer_class in selected_tokenizers:
        # Look up the name of the tokenizer
        tokenizer_name = next((x[0] for x in tokenizers_list if x[1] == tokenizer_class), None)
        if not tokenizer_name:
            continue  # Skip if the tokenizer name is not found

        tokenizer = tokenizer_class()
        tokenized_file_folder = os.path.join(tokenization_dir, f"tokenizer_{tokenizer_name}")
        os.mkdir(tokenized_file_folder)
        for file in tqdm(tokenizable_files, ncols=70):  # Wrap tokenizable_files with tqdm for progress bar
            try:
                midi_file_path = os.path.join(folder_path, file)
                tokens = tokenizer(midi_file_path)
                tokens_string = str(tokens)

                # check if the tokens are in a nested list
                if "[[" in tokens_string and "]]" in tokens_string:
                    tokens_list = extract_tokens_nested_list(tokens_string)
                else:
                    tokens_list = extract_tokens(tokens_string)

                # Convert the list of tokens to a DataFrame
                df = pd.DataFrame(tokens_list)

                # Get the column names for the current tokenizer
                headers = tokenizer_headers.get(tokenizer_name, [])

                # Check if the DataFrame has the right number of columns
                if len(df.columns) == len(headers):
                    df.columns = headers
                else:
                    print(
                        f'Warning: Number of columns in the DataFrame does not match the number of headers for tokenizer {tokenizer_name}. Defaulting to generic headers.')
                    df.columns = ['column' + str(i) for i in range(1, len(df.columns) + 1)]

                # Save the DataFrame to a CSV file
                base_name = os.path.splitext(file)[0]  # Get file base name without extension
                file_name_token = f"{base_name}_tokenizer_{tokenizer_name}_tokens.csv"
                file_path = os.path.join(tokenized_file_folder, file_name_token)

                df.to_csv(file_path, index=False)
                num_tokenized += 1
                files_status.append([tokenizer_name, file, "<successfully converted>"])

            except Exception as e:
                files_status.append([tokenizer_name, file, str(e)])
                list = [tokenizer_name, os.path.join(tokenized_file_folder, file), str(e)]
                create_log_entry(list, tokenization_dir + r'\tokenization_log.xlsx')

        combine_csv_files_in_directory(tokenized_file_folder, f"combined_tokenizer_{tokenizer_name}_tokens.csv")

    # Create text_dict with both tokenization success rate and failed files
    failed_files_status = [status for status in files_status if "<successfully converted>" not in status]

    if failed_files_status:  # If there are failed tokenizations
        # Filter out only the failed tokenizations and create text_dict with failed files only
        text_dict = {
            "menu_displayed_text": [
                "File Tokenization: Tokenization Summary",
                f"Conversion Success Rate {num_tokenized}/{num_files_tokenized} ({(num_tokenized / num_files_tokenized) * 100:.2f}%)",
                "<To continue, please press Enter>",
                ["Tokenizer name", "File name", "Details"]
            ],
            "menu_entries_results": [
                [str(tokenizer_name), str(file), f"'{str(status)}'"] for tokenizer_name, file, status in
                failed_files_status]
        }
    else:  # If there are no failed tokenizations
        # Show the first 30 successful tokenizations, or all if less than 30
        num_successful_files = min(30, len(files_status))
        successful_files_status = files_status[:num_successful_files]
        successful_files_display = successful_files_status
        if len(files_status) > 30:  # If there are more than 30 files, add an indicator at the end
            successful_files_display.append(["...", "...", "<more files not shown>"])
        text_dict = {
            "menu_displayed_text": [
                "File Tokenization: Tokenization Summary",
                f"Conversion Success Rate {num_tokenized}/{num_files_tokenized} ({(num_tokenized / num_files_tokenized) * 100:.2f}%)",
                "<To continue, please press Enter>",
                ["Tokenizer name", "File name", "Details"]
            ],
            "menu_entries_results": [
                [str(tokenizer_name), str(file), f"'{str(status)}'"] for tokenizer_name, file, status in
                successful_files_display]
        }

    display_menu_print_results(text_dict)


def combine_csv_files_in_directory(directory_path, output_file_name):
    """
    Combine all CSV files in a given directory into a single CSV file.

    Parameters
    ----------
    directory_path : str
        Path to the directory with the CSV files.
    output_file_name : str
        Name of the combined output CSV file.

    Returns
    -------
    None
    """

    # Get a list of all CSV files in the directory
    csv_files = glob.glob(os.path.join(directory_path, "*.csv"))

    # Initialize an empty list to hold DataFrames
    df_list = []

    print("\nCombining CSV files into one...")

    # Loop through each CSV file
    for csv_file in tqdm(csv_files, ncols=70):  # Wrap csv_files with tqdm for progress bar
        # Read the CSV file into a DataFrame
        df = pd.read_csv(csv_file)

        # Extract the original file name from the CSV file name
        original_file_name = os.path.basename(csv_file).split('_tokenizer')[0]
        # Assuming file names are in the format "<original_name>_tokenizer_<tokenizer_name>_tokens.csv"

        # Insert it into a new column at the start of the DataFrame
        df.insert(0, 'filename', original_file_name)

        # Append the DataFrame to the list
        df_list.append(df)

    # Concatenate all DataFrames in the list
    combined_df = pd.concat(df_list, ignore_index=True)

    # Save the combined DataFrame to a CSV file
    combined_df.to_csv(os.path.join(directory_path, output_file_name), index=False)


def list_tokenizable_files(folder_path):
    """
    Lists all MIDI files in a given folder.

    Parameters
    ----------
    folder_path : str
        Path to the folder to be searched.

    Returns
    -------
    list
        List of MIDI file names.
    """
    tokenizable_extensions = ['.midi', '.mid']
    tokenizable_files = [f for f in os.listdir(folder_path) if os.path.splitext(f)[1].lower() in tokenizable_extensions]
    return tokenizable_files


def select_tokenizer():
    """
    Provides a user interface for tokenizer selection.

    The user can select one or more tokenizers from a predefined list.

    Returns
    -------
    list
        List of selected tokenizer classes.
    """
    tokenizer_dict = {
        "menu_displayed_text": [
            "File Tokenization: Tokenizer Selection",
            "Please select one of the following tokenizers by entering the corresponding index number:",
            "Which tokenizer do you want to select? (<No. of menu item>): ",
            ["Tokenizer", "Description"]
        ],
        "menu_entries": [[tokenizer[0], tokenizer[1], tokenizer[2]] for tokenizer in tokenizers_list] + [
            ["All", "return_all", "Apply all tokenizers"]]
    }
    choice = display_menu_request_selection(tokenizer_dict)

    if choice == "return_all":
        return [tokenizer[1] for tokenizer in tokenizers_list]
    else:
        return [choice]


def select_folder():
    """
    User interface for folder selection.

    The user can input the path to the desired folder. If the path does not exist or
    does not lead to a directory, an error message is displayed.

    Returns
    -------
    str
        Path to the selected folder.
    """
    while True:
        folder_path_dict = {
            "menu_displayed_text": [
                "File Tokenization: Folder Selection",
                "Please enter the path to the folder of music files. You can also drag and drop the folder into this terminal: ",
                "Enter folder path here or drag and drop folder: ",
                ["Menu item", "<Explanation>"]
            ],
            "menu_entries_text": [
                ["Tips:", "1. You can copy the path from your file explorer and paste it here."],
                ["",
                 "2. If your terminal supports it, you can also drag and drop the folder directly into the terminal."]
            ]
        }
        folder_path = display_menu_print_textblock(folder_path_dict)

        if os.path.isdir(folder_path):
            if len(list_tokenizable_files(folder_path)) > 0:
                return folder_path
            else:
                error_message_dict = {
                    "menu_displayed_text": [
                        "File Conversion: Folder Selection - Error",
                        "No MIDI (.midi or .mid) files in the directory.",
                        "<To continue, please press Enter>",
                        ["", "Troubleshooting assistance:"]
                    ],
                    "menu_entries_text": [
                        ["Reason 1:", "There are no MIDI (.midi or .mid) files in the selected directory."],
                        ["Reason 2:", "Make sure your files have the correct extensions (.midi or .mid)."],
                        ["Reason 3:",
                         "Ensure that the selected directory is not protected or read-only."]
                    ]
                }
                display_menu_print_textblock(error_message_dict)
        else:
            error_message_dict = {
                "menu_displayed_text": [
                    "File Conversion: Folder Selection - Error",
                    "Invalid path. Please check the following possible reasons:",
                    "<To continue, please press Enter>",
                    ["", "Troubleshooting assistance:"]
                ],
                "menu_entries_text": [
                    ["Reason 1:", "The path does not exist."],
                    ["Reason 2:", "The path does not lead to a directory."],
                    ["Reason 3:",
                     "The path includes special characters that are not supported by your operating system."]
                ]
            }
            display_menu_print_textblock(error_message_dict)


def extract_tokens(tokens_string):
    """
    Extracts tokens from a string representation of a list.

    Parameters
    ----------
    tokens_string : str
        String representation of a list of tokens.

    Returns
    -------
    list
        List of tokens.
    """
    # cut away the beginning of the string
    start_index = tokens_string.index('=[') + 2
    trimmed_string = tokens_string[start_index:]

    # cut away everything after the first appearing "]"
    end_index = trimmed_string.index(']')
    trimmed_string = trimmed_string[:end_index]

    # extract strings using regular expression
    search_pattern = r"'([^']*)'"
    token_list = re.findall(search_pattern, trimmed_string)

    return token_list


def extract_tokens_nested_list(tokens_string):
    """
    Extracts tokens from a string representation of a nested list.

    Parameters
    ----------
    tokens_string : str
        String representation of a nested list of tokens.

    Returns
    -------
    list
        List of token lists.
    """
    # cut away everything before the string "[TokSequence(tokens=["
    start_index = tokens_string.index('=[') + 2
    trimmed_string = tokens_string[start_index:]

    # cut away everything after the string ", ids="
    end_index = trimmed_string.index('], ids=')
    trimmed_string = trimmed_string[:end_index]

    search_pattern = r'\[([^\[\]]*?)\]'
    matched_lists = re.findall(search_pattern, trimmed_string)

    token_lists = [ast.literal_eval(list) for list in matched_lists]

    return token_lists


def tokenize_midi_file(midi_file_path, tokenizer_class, tokenized_file_folder):
    """
    Tokenizes a MIDI file using a specified tokenizer class and saves the tokenized sequence to a
    CSV file.

    Parameters
    ----------
    midi_file_path : str
        Path to the MIDI file to be tokenized.
    tokenizer_class : class
        Class of the tokenizer to be used.
    tokenized_file_folder : str
        Path to the folder where the CSV file will be saved.

    Returns
    -------
    None
    """
    # create the tokenizer and convert the MIDI to tokens
    tokenizer = tokenizer_class()
    tokens = tokenizer(midi_file_path)

    tokens_string = str(tokens)

    # check if the tokens are in a nested list
    if "[[" in tokens_string and "]]" in tokens_string:
        tokens_list = extract_tokens_nested_list(tokens_string)
    else:
        tokens_list = extract_tokens(tokens_string)

    # convert the list of tokens to a DataFrame
    df = pd.DataFrame(tokens_list)

    # save the DataFrame to a CSV file
    file_name_token = f"{tokenizer_class.__name__}_tokens.csv"
    file_path = os.path.join(tokenized_file_folder, file_name_token)

    df.to_csv(file_path, index=False, header=False)

    # print(f"Tokens saved to {file_path}")


def create_log_entry(list, list_path):
    """
    Appends a new entry to an existing .xlsx file.

    If the file does not exist, it is created.

    Parameters
    ----------
    list : list
        List of strings to be appended as a new row.
    list_path : str
        Path to the .xlsx file.

    Returns
    -------
    None
    """
    wb = load_workbook(list_path)
    ws = wb.active
    ws.append(list)
    wb.save(list_path)


corpus_tokenization()
